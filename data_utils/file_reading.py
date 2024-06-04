import pandas as pd
import os
import openpyxl


def read_folder_and_get_union(datadir, filesep=',', staic_col_params=[]):
    '''

    Args:
        datadir:
        filesep:
        staic_col_params:

    Returns:
    staic_col_params : [
        {
            'colname': 'date',
            'colvalue': '2022-01-01'
        }

    ]
    '''
    datafiles = os.listdir(datadir)

    datadfs = []
    for _file in datafiles:
        _path = os.path.join(datadir, _file)
        _df = pd.read_csv(_path, sep=filesep)
        datadfs.append(_df)

    if len(datadfs) == 0:
        return None

    df_merged = pd.concat(datadfs, ignore_index=True)

    for _item in staic_col_params:
        df_merged[_item['colname']] = _item['colvalue']

    return df_merged


def get_excel_sheet_names(filepath):
    brand_file = openpyxl.load_workbook(filepath)
    sheet_names = brand_file.sheetnames
    return sheet_names


def read_excel_sheet_data(file_path, sheet_name, has_header=True):
    # Load the workbook and select the specified sheet
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet name '{sheet_name}' does not exist in the workbook.")

    sheet = workbook[sheet_name]

    # Extract data into a list
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    if has_header:
        data = data[1:]

    return data


